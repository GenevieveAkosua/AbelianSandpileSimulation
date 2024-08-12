# Adapted from: Lord - Full Stack [https://www.youtube.com/watch?v=t9iDpwl6fMQ]

from PIL import Image
import numpy as np
import openpyxl
import imagehash
import pandas as panda
import subprocess
import os


def run_simulation(input_file, output_file, version_folder):
    original_dir = os.getcwd()
    os.chdir(version_folder)

    command = f'make run ARGS="input/{input_file} output/{output_file}"'
    result = subprocess.run(command, capture_output=True, text=True, shell=True)

    os.chdir(original_dir)
    return result.stdout.strip().split('\n')


def extract_metrics(output_lines):
    steps_line = next((line for line in output_lines if "Number of steps to stable state" in line), None)
    time_line = next((line for line in output_lines if "Time:" in line), None)

    steps = int(steps_line.split(":")[1].strip()) if steps_line else None
    time = int(time_line.split(":")[1].strip().split()[0]) if time_line else None

    return steps, time


def compare_outputs(serial_fol, parallel_fol, output_file):
    # Create file paths for serial and parallel folders
    serial_path = os.path.join(serial_fol, 'output', output_file)
    parallel_path = os.path.join(parallel_fol, 'output', output_file)

    # Open the images
    serial_image = Image.open(serial_path)
    parallel_image = Image.open(parallel_path)

    # Get the average hash value
    hash_one = imagehash.average_hash(serial_image)
    hash_two = imagehash.average_hash(parallel_image)

    difference = hash_one - hash_two

    return difference == 0


# Specify variables for the trials
TRIALS = 10
serial_folder = "/Users/genevievechikwanha/Library/CloudStorage/OneDrive-UniversityofCapeTown/Uni Work/Computer Science/CSC2002S/CSC2002S 2024/1. Paralell Programming/Assignments/Assignment1/PCP_ParallelAssignment2024"
parallel_folder = "/Users/genevievechikwanha/Library/CloudStorage/OneDrive-UniversityofCapeTown/Uni Work/Computer Science/CSC2002S/CSC2002S 2024/1. Paralell Programming/Assignments/Assignment1/CHKKAR002"
io_pairs = [
    ('32_by_32_all_8.csv', '32.png'),
    ('125_by_125_center_1111.csv', '125_center.png'),
    ('750_by_750_center_777.csv', '750_777.png')
]

# ('8_by_8_all_4.csv', '8.png'),
# ('16_by_16_all_4.csv', '16.png'),
# ('16_by_16_one_100.csv', '16One.png'),
# ('65_by_65_all_4.csv', '65.png'),
# ('125_by_125_all_0.csv', '125.png'),
# ('250_by_250_all_4.csv', '250.png'),
# ('517_by_517_centre_534578.csv', '517.png'),
# ('750_by_750_all_0.csv', '750.png'),
# ('1001_by_1001_all_8.csv', '1001.png')

all_results = []

for input_file, output_file in io_pairs:
    trial_results = []
    for trial in range(TRIALS):
        # Run serial version
        serial_output = run_simulation(input_file, output_file, serial_folder)
        serial_steps, serial_time = extract_metrics(serial_output)

        # Run parallel version
        parallel_output = run_simulation(input_file, output_file, parallel_folder)
        parallel_steps, parallel_time = extract_metrics(parallel_output)

        # Store results for current trial
        trial_results.append({
            'input_file': input_file,
            'output_file': output_file,
            'trial': trial,
            'serial_timesteps': serial_steps,
            'serial_timetaken': serial_time,
            'parallel_timesteps': parallel_steps,
            'parallel_timetaken': parallel_time
        })

    # Check correctness after all trials
    is_correct = compare_outputs(serial_folder, parallel_folder, output_file)

    # Add similarity to all trial results for this input/output pair
    for result in trial_results:
        result['image_similarity'] = is_correct

    all_results.extend(trial_results)

    # Add correctness score to all trial results for this input/output pair
    # for result in all_results:
        #if result['input_file'] == input_file and result['output_file'] == output_file:
            #result['image_correctness'] = is_correct

    print(f"Completed {TRIALS} trials for {input_file} -> {output_file}")

# Calculate minimum values and correctness for each io pair
min_results = {}
for result in all_results:
    input_file, output_file = result['input_file'], result['output_file']
    if (input_file, output_file) not in min_results:
        min_results[(input_file, output_file)] = {
            'input_file': input_file,
            'output_file': output_file,
            'serial_timesteps': float('inf'),
            'serial_timetaken': float('inf'),
            'parallel_timesteps': float('inf'),
            'parallel_timetaken': float('inf'),
            'correctness_rate': 0,
            'count': 0
        }

    min_results[(input_file, output_file)]['serial_timesteps'] = min(
        min_results[(input_file, output_file)]['serial_timesteps'], result['serial_timesteps'])
    min_results[(input_file, output_file)]['serial_timetaken'] = min(
        min_results[(input_file, output_file)]['serial_timetaken'], result['serial_timetaken'])
    min_results[(input_file, output_file)]['parallel_timesteps'] = min(
        min_results[(input_file, output_file)]['parallel_timesteps'], result['parallel_timesteps'])
    min_results[(input_file, output_file)]['parallel_timetaken'] = min(
        min_results[(input_file, output_file)]['parallel_timetaken'], result['parallel_timetaken'])
    min_results[(input_file, output_file)]['correctness_rate'] += result['image_similarity']
    min_results[(input_file, output_file)]['count'] += 1

min_results = [v for v in min_results.values()]
min_results = sorted(min_results, key=lambda x: (x['input_file'], x['output_file']))

# Write results to Excel using openpyxl
workbook = openpyxl.Workbook()

# Sheet for all data
worksheet_all = workbook.active
worksheet_all.title = "All Data"

# Write header row for all data
header_all = ['input_file', 'output_file', 'trial', 'serial_timesteps', 'serial_timetaken', 'parallel_timesteps', 'parallel_timetaken', 'image_similarity']
for col_idx, value in enumerate(header_all):
    worksheet_all.cell(row=1, column=col_idx+1, value=value)

# Write all data
for row_idx, result in enumerate(all_results):
    row = [result['input_file'], result['output_file'], result['trial'], result['serial_timesteps'], result['serial_timetaken'], result['parallel_timesteps'], result['parallel_timetaken'], result['image_similarity']]
    for col_idx, value in enumerate(row):
        worksheet_all.cell(row=row_idx+2, column=col_idx+1, value=value)

# Sheet for minimum values
worksheet_min = workbook.create_sheet(title="Minimum Values")

# Write header row for minimum values
header_min = ['input_file', 'output_file', 'serial_timesteps', 'serial_timetaken', 'parallel_timesteps', 'parallel_timetaken', 'correctness_rate']
for col_idx, value in enumerate(header_min):
    worksheet_min.cell(row=1, column=col_idx+1, value=value)

# Write minimum values
for row_idx, result in enumerate(min_results):
    row = [result['input_file'], result['output_file'], result['serial_timesteps'], result['serial_timetaken'], result['parallel_timesteps'], result['parallel_timetaken'], result['correctness_rate'] / result['count']]
    for col_idx, value in enumerate(row):
        worksheet_min.cell(row=row_idx+2, column=col_idx+1, value=value)

workbook.save('Results.xlsx')

print("All trials are complete and results have been saved to results.xlsx")
print("\nSummary:")
for result in min_results:
    print(f"Input: {result['input_file']}, Output: {result['output_file']}")
    print(f"  Serial timesteps: {result['serial_timesteps']}")
    print(f"  Serial timetaken: {result['serial_timetaken']}")
    print(f"  Parallel timesteps: {result['parallel_timesteps']}")
    print(f"  Parallel timetaken: {result['parallel_timetaken']}")
    print(f"  Correctness rate: {result['correctness_rate'] / result['count']:.2f}")
    print()
